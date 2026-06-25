"""Leitor da fatura de cartão do Itaú exportada em planilha (.xlsx).

O Itaú não permite exportar a fatura do cartão em OFX — só em planilha. Este
módulo lê o XLSX e devolve os mesmos `OfxTransaction` usados pelo importador de
OFX, para reaproveitar de graça a dedup por impressão digital e a categorização.

Convenção de sinais: na fatura, compra vem positiva e pagamento negativo. O app
usa negativo = saída, então invertemos o sinal (compra → negativo, pagamento →
positivo, virando um crédito que abate a fatura).
"""
import io
import re
import unicodedata
from datetime import date, datetime

from openpyxl import load_workbook

from ofx_import import OfxTransaction


def _norm(value) -> str:
    """Texto da célula sem acento, minúsculo e sem espaços nas bordas."""
    if value is None:
        return ""
    text = unicodedata.normalize("NFKD", str(value)).encode("ascii", "ignore").decode()
    return text.strip().lower()


_MONTHS = {
    "janeiro": 1, "fevereiro": 2, "marco": 3, "abril": 4, "maio": 5, "junho": 6,
    "julho": 7, "agosto": 8, "setembro": 9, "outubro": 10, "novembro": 11, "dezembro": 12,
}


def _competencia(rows, header_idx) -> date | None:
    """Mês em que esta fatura é descontada. Prioriza a data de Vencimento (célula
    de data concreta); se não houver, lê o mês/ano do título "Fatura ... Mês/Ano"."""
    # 1) Data de vencimento: acha o rótulo e pega a primeira data na mesma coluna.
    for i in range(header_idx):
        for j, c in enumerate(rows[i]):
            if _norm(c) == "vencimento":
                for k in range(i, min(i + 4, header_idx)):
                    cell = rows[k][j] if j < len(rows[k]) else None
                    if isinstance(cell, (date, datetime)):
                        return cell.date() if isinstance(cell, datetime) else cell
    # 2) Título da fatura.
    text = _norm(" ".join(str(c) for r in rows[:header_idx] for c in r if c is not None))
    m = re.search(r"fatura.*?(" + "|".join(_MONTHS) + r")\D*(\d{4})", text)
    if m:
        return date(int(m.group(2)), _MONTHS[m.group(1)], 1)
    return None


# Rótulos aceitos para cada coluna da seção de lançamentos.
_COL_ALIASES = {
    "date": {"data"},
    "desc": {"lancamento", "descricao", "estabelecimento"},
    "inst": {"parcelamento"},
    "amount": {"valor"},
}


def parse_card_xlsx(raw_bytes: bytes) -> list[OfxTransaction]:
    wb = load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # 1) Localiza a linha de cabeçalho da tabela de lançamentos (tem Data e Valor).
    cols: dict[str, int] = {}
    header_idx = None
    for i, row in enumerate(rows):
        labels = [_norm(c) for c in row]
        if "data" in labels and "valor" in labels:
            for j, lab in enumerate(labels):
                for key, aliases in _COL_ALIASES.items():
                    if lab in aliases:
                        cols.setdefault(key, j)
            header_idx = i
            break

    if header_idx is None or "date" not in cols or "amount" not in cols:
        raise ValueError(
            "Planilha não reconhecida. Esperava uma fatura de cartão do Itaú "
            "com colunas Data, Lançamento e Valor."
        )

    competencia = _competencia(rows, header_idx)

    # 2) Lê cada linha de dados até o rodapé (linhas sem data/valor válidos são puladas).
    txns: list[OfxTransaction] = []
    for row in rows[header_idx + 1:]:
        d = row[cols["date"]] if cols["date"] < len(row) else None
        v = row[cols["amount"]] if cols["amount"] < len(row) else None
        if not isinstance(d, (date, datetime)) or v is None:
            continue
        try:
            value = float(v)
        except (TypeError, ValueError):
            continue

        bought_on = d.date() if isinstance(d, datetime) else d
        desc = ""
        if "desc" in cols and cols["desc"] < len(row) and row[cols["desc"]]:
            desc = str(row[cols["desc"]]).strip()
        desc = desc or "Sem descrição"

        # Parcela: entra na descrição (cada parcela é uma cobrança distinta) e, para
        # orçamentos/estatísticas, é lançada no mês em que é DESCONTADA (a competência
        # da fatura), não na data da compra original — que fica preservada no memo.
        instalment = None
        if "inst" in cols and cols["inst"] < len(row) and row[cols["inst"]]:
            instalment = str(row[cols["inst"]]).strip()
            desc = f"{desc} ({instalment})"

        if instalment and competencia:
            posted = competencia
            raw_memo = f"{desc} | compra em {bought_on.isoformat()}"
        else:
            posted = bought_on
            raw_memo = desc

        amount = -value  # compra (positiva) → saída (negativa); pagamento → crédito
        txns.append(OfxTransaction(
            fitid=None,
            posted_on=posted.isoformat(),
            amount=amount,
            description=desc,
            raw_memo=raw_memo,
            trntype="DEBIT" if amount < 0 else "CREDIT",
        ))
    return txns


def read_card_xlsx_file(raw_bytes: bytes) -> list[OfxTransaction]:
    """Alias simétrico ao read_ofx_file, para o dispatch no app."""
    return parse_card_xlsx(raw_bytes)
