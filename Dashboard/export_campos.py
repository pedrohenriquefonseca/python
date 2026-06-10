"""
Exporta campos do PWA para Excel com 3 abas:
  - Projeto: campos do objeto projeto
  - Tarefa Resumo: campos da tarefa de resumo do projeto
  - Tarefa Normal: campos de uma tarefa individual
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
import pwa_client


def flatten(obj, prefix=""):
    rows = []
    if isinstance(obj, dict):
        for k, v in obj.items():
            if k == "__metadata":
                continue
            key = f"{prefix}.{k}" if prefix else k
            if isinstance(v, dict):
                if "deferred" in v:
                    rows.append((key, "[navigation - nao expandido]"))
                else:
                    rows.extend(flatten(v, key))
            elif isinstance(v, list):
                rows.append((key, f"[lista com {len(v)} item(s)]"))
            else:
                rows.append((key, v))
    return rows


def make_sheet(ws, rows):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    headers = ["Campo", "Valor (exemplo)", "Usar no dashboard?", "Anotacao"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1e3a5f")
        cell.alignment = Alignment(horizontal="center")

    for i, (campo, valor) in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=campo)
        ws.cell(row=i, column=2, value=str(valor) if valor is not None else "")
        if i % 2 == 0:
            for col in range(1, 5):
                ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor="f0f4f8")

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 40
    ws.freeze_panes = "A2"


def main():
    import openpyxl

    print("Buscando projetos...")
    projects = pwa_client._get_all(
        f"{pwa_client.PS_BASE}/Projects",
        {"$top": "1", "$expand": "ProjectSummaryTask,EnterpriseProjectType,Owner,Phase,Stage"},
    )

    if not projects:
        print("Nenhum projeto encontrado.")
        return

    p   = projects[0]
    pid = p.get("Id")
    print(f"Projeto: {p.get('Name')}  ({pid})")

    # Aba 1 — campos do projeto
    proj_rows = flatten(p)

    # Aba 2 e 3 — campos das tarefas
    # Busca todos os projetos para achar um com tarefas publicadas
    print("Buscando todos os projetos para achar um com tarefas...")
    all_projects = pwa_client._get_all(
        f"{pwa_client.PS_BASE}/Projects",
        {"$top": "200"},
    )

    tasks = []
    proj_com_tarefas = None
    for candidate in all_projects:
        cid = candidate.get("Id")
        try:
            t = pwa_client._get_all(
                f"{pwa_client.PS_BASE}/Projects('{cid}')/Tasks",
                {"$top": "5"},
            )
            if t:
                tasks = t
                proj_com_tarefas = candidate
                print(f"  Tarefas encontradas no projeto: {candidate.get('Name')}")
                break
        except Exception:
            continue

    if not tasks:
        print("  Nenhum projeto com tarefas acessiveis encontrado.")
    else:
        print(f"  {len(tasks)} tarefa(s) de amostra.")

    summary = next((t for t in tasks if t.get("IsProjectSummary")), tasks[0] if tasks else {})
    normal  = next((t for t in tasks if not t.get("IsProjectSummary") and not t.get("Summary")),
                   tasks[-1] if len(tasks) > 1 else {})

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Projeto"
    make_sheet(ws1, proj_rows)

    ws2 = wb.create_sheet("Tarefa Resumo")
    make_sheet(ws2, flatten(summary) if summary else [])

    ws3 = wb.create_sheet("Tarefa Normal")
    make_sheet(ws3, flatten(normal) if normal else [])

    output = Path(__file__).parent / "campos_projeto.xlsx"
    wb.save(output)
    print(f"Arquivo salvo em: {output}")


if __name__ == "__main__":
    main()
